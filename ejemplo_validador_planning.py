"""
═══════════════════════════════════════════════════════════════
EJEMPLO DE AUTOMATIZACIÓN - VALIDACIÓN DE PLANNINGS
Sistema desarrollado por: Julián Alexander Juárez Alvarado
CEDIS Cancún 427 - Tiendas Chedraui
═══════════════════════════════════════════════════════════════

Este módulo demuestra las capacidades de automatización mediante:
- Conexión a IBM DB2 (Manhattan WMS)
- Procesamiento de datos con Pandas
- Validaciones inteligentes
- Generación de reportes Excel profesionales
- Sistema de alertas por criticidad
"""

import os
import pandas as pd
import pyodbc
from datetime import datetime
from typing import Dict, List
import logging

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ValidadorPlanning:
    """
    Sistema de validación automatizada de plannings.
    
    Reemplaza proceso manual de 2-3 horas por validación
    automática de 15-20 minutos (85-90% reducción).
    """
    
    def __init__(self, db_config: Dict[str, str]):
        """
        Inicializa el validador con configuración de BD.
        
        Args:
            db_config: Diccionario con credenciales DB2
        """
        self.db_config = db_config
        self.connection = None
        self.resultados = []
        
    def conectar_db2(self) -> bool:
        """
        Establece conexión con base de datos IBM DB2.
        
        Returns:
            bool: True si conexión exitosa, False en caso contrario
        """
        try:
            connection_string = (
                f"DRIVER={{IBM DB2 ODBC DRIVER}};"
                f"DATABASE={self.db_config['database']};"
                f"HOSTNAME={self.db_config['host']};"
                f"PORT={self.db_config['port']};"
                f"PROTOCOL=TCPIP;"
                f"UID={self.db_config['user']};"
                f"PWD={self.db_config['password']};"
            )
            
            self.connection = pyodbc.connect(connection_string)
            logger.info("✅ Conexión a DB2 establecida")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error conectando a DB2: {e}")
            return False
    
    def validar_oc_vs_distribuciones(self, orden_compra: str) -> Dict:
        """
        Valida cuadre entre Orden de Compra y Distribuciones.
        
        Args:
            orden_compra: Número de OC a validar
            
        Returns:
            Dict con resultados de validación y alertas
        """
        query_oc = f"""
        SELECT 
            AHTLOC AS ALMACEN,
            AHPON AS OC,
            AHVNAM AS PROVEEDOR,
            AHITMN AS SKU,
            AHACTO AS CANTIDAD_OC,
            AHASTP AS STATUS_ASN
        FROM WM260BASD.AHASNF00
        WHERE AHPON = '{orden_compra}'
          AND AHTLOC = '427'
          AND AHASTP = '1'
        """
        
        query_distros = f"""
        SELECT 
            DSTLOC AS ALMACEN,
            DSPON AS OC,
            DSITMN AS SKU,
            SUM(DSDQTY) AS CANTIDAD_DISTRO
        FROM WM260BASD.DSTRBF00
        WHERE DSPON = '{orden_compra}'
          AND DSTLOC = '427'
        GROUP BY DSTLOC, DSPON, DSITMN
        """
        
        try:
            # Ejecutar queries
            df_oc = pd.read_sql(query_oc, self.connection)
            df_distros = pd.read_sql(query_distros, self.connection)
            
            # Merge y comparación
            df_merged = pd.merge(
                df_oc,
                df_distros,
                left_on=['ALMACEN', 'OC', 'SKU'],
                right_on=['ALMACEN', 'OC', 'SKU'],
                how='outer',
                indicator=True
            )
            
            # Análisis de resultados
            total_ocs = len(df_oc)
            total_distros = len(df_distros)
            cuadres_ok = len(df_merged[
                (df_merged['CANTIDAD_OC'] == df_merged['CANTIDAD_DISTRO']) &
                (df_merged['_merge'] == 'both')
            ])
            
            diferencias = df_merged[
                (df_merged['CANTIDAD_OC'] != df_merged['CANTIDAD_DISTRO']) |
                (df_merged['_merge'] != 'both')
            ]
            
            # Clasificar alertas por criticidad
            alertas = self._clasificar_alertas(diferencias)
            
            resultado = {
                'oc': orden_compra,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total_skus_oc': total_ocs,
                'total_skus_distro': total_distros,
                'cuadres_ok': cuadres_ok,
                'porcentaje_cuadre': (cuadres_ok / total_ocs * 100) if total_ocs > 0 else 0,
                'alertas': alertas,
                'status': self._determinar_status(alertas)
            }
            
            logger.info(f"✅ OC {orden_compra} validada: {resultado['porcentaje_cuadre']:.1f}% cuadre")
            return resultado
            
        except Exception as e:
            logger.error(f"❌ Error validando OC {orden_compra}: {e}")
            return {
                'oc': orden_compra,
                'error': str(e),
                'status': 'ERROR'
            }
    
    def _clasificar_alertas(self, diferencias: pd.DataFrame) -> List[Dict]:
        """
        Clasifica diferencias en alertas por nivel de criticidad.
        
        Niveles:
        - CRÍTICA: SKU en OC sin distribuciones
        - MEDIA: Diferencia de cantidades
        - BAJA: Distribuciones sin SKU en OC
        """
        alertas = []
        
        # Alertas CRÍTICAS: SKU sin distribuciones
        sin_distro = diferencias[diferencias['_merge'] == 'left_only']
        for _, row in sin_distro.iterrows():
            alertas.append({
                'nivel': 'CRÍTICA',
                'tipo': 'SKU_SIN_DISTRIBUCION',
                'sku': row['SKU'],
                'cantidad_oc': row['CANTIDAD_OC'],
                'mensaje': f"SKU {row['SKU']} tiene {row['CANTIDAD_OC']} unidades en OC pero NO tiene distribuciones"
            })
        
        # Alertas MEDIAS: Diferencia de cantidades
        diff_cantidades = diferencias[
            (diferencias['_merge'] == 'both') &
            (diferencias['CANTIDAD_OC'] != diferencias['CANTIDAD_DISTRO'])
        ]
        for _, row in diff_cantidades.iterrows():
            alertas.append({
                'nivel': 'MEDIA',
                'tipo': 'DIFERENCIA_CANTIDAD',
                'sku': row['SKU'],
                'cantidad_oc': row['CANTIDAD_OC'],
                'cantidad_distro': row['CANTIDAD_DISTRO'],
                'diferencia': abs(row['CANTIDAD_OC'] - row['CANTIDAD_DISTRO']),
                'mensaje': f"SKU {row['SKU']}: OC={row['CANTIDAD_OC']}, Distros={row['CANTIDAD_DISTRO']}"
            })
        
        # Alertas BAJAS: Distribuciones sin OC
        distro_sin_oc = diferencias[diferencias['_merge'] == 'right_only']
        for _, row in distro_sin_oc.iterrows():
            alertas.append({
                'nivel': 'BAJA',
                'tipo': 'DISTRO_SIN_OC',
                'sku': row['SKU'],
                'cantidad_distro': row['CANTIDAD_DISTRO'],
                'mensaje': f"SKU {row['SKU']} tiene distribuciones pero NO está en OC"
            })
        
        return alertas
    
    def _determinar_status(self, alertas: List[Dict]) -> str:
        """
        Determina status general basado en alertas.
        """
        if not alertas:
            return '✅ OK'
        
        medias = 0
        for a in alertas:
            nivel = a.get('nivel')
            if nivel == 'CRÍTICA':
                return '🔴 CRÍTICO'
            if nivel == 'MEDIA':
                medias += 1

        if medias > 3:
            return '🟡 REVISAR'
        else:
            return '🟢 ACEPTABLE'
    
    def validar_multiples_ocs(self, lista_ocs: List[str]) -> List[Dict]:
        """
        Valida múltiples OCs en batch.
        
        Args:
            lista_ocs: Lista de números de OC
            
        Returns:
            Lista de resultados de validación
        """
        logger.info(f"📋 Iniciando validación de {len(lista_ocs)} OCs...")
        
        resultados = []
        for i, oc in enumerate(lista_ocs, 1):
            logger.info(f"  [{i}/{len(lista_ocs)}] Validando OC {oc}...")
            resultado = self.validar_oc_vs_distribuciones(oc)
            resultados.append(resultado)
        
        logger.info(f"✅ Validación completada: {len(resultados)} OCs procesadas")
        return resultados
    
    def generar_reporte_excel(self, resultados: List[Dict], archivo_salida: str):
        """
        Genera reporte Excel profesional con formato corporativo Chedraui.
        
        Args:
            resultados: Lista de resultados de validación
            archivo_salida: Ruta del archivo Excel a generar
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        logger.info(f"📊 Generando reporte Excel: {archivo_salida}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Validación Planning"
        
        # Colores corporativos Chedraui
        COLOR_AZUL = "0066B3"
        COLOR_NARANJA = "FF6600"
        COLOR_VERDE = "10B981"
        COLOR_AMARILLO = "F59E0B"
        COLOR_ROJO = "EF4444"
        
        # Encabezado principal
        ws.merge_cells('A1:G1')
        ws['A1'] = "REPORTE DE VALIDACIÓN DE PLANNINGS"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=COLOR_AZUL, end_color=COLOR_AZUL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Información general
        ws['A2'] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Total OCs Validadas: {len(resultados)}"
        
        # Encabezados de tabla
        headers = ['OC', 'SKUs OC', 'SKUs Distro', 'Cuadres OK', '% Cuadre', 'Alertas', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_NARANJA, end_color=COLOR_NARANJA, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        row = 6
        for resultado in resultados:
            ws.cell(row=row, column=1, value=resultado['oc'])
            ws.cell(row=row, column=2, value=resultado.get('total_skus_oc', 0))
            ws.cell(row=row, column=3, value=resultado.get('total_skus_distro', 0))
            ws.cell(row=row, column=4, value=resultado.get('cuadres_ok', 0))
            ws.cell(row=row, column=5, value=f"{resultado.get('porcentaje_cuadre', 0):.1f}%")
            ws.cell(row=row, column=6, value=len(resultado.get('alertas', [])))
            ws.cell(row=row, column=7, value=resultado.get('status', 'N/A'))
            
            # Colorear status
            status_cell = ws.cell(row=row, column=7)
            if '✅' in resultado.get('status', ''):
                status_cell.fill = PatternFill(start_color=COLOR_VERDE, end_color=COLOR_VERDE, fill_type="solid")
            elif '🔴' in resultado.get('status', ''):
                status_cell.fill = PatternFill(start_color=COLOR_ROJO, end_color=COLOR_ROJO, fill_type="solid")
            elif '🟡' in resultado.get('status', ''):
                status_cell.fill = PatternFill(start_color=COLOR_AMARILLO, end_color=COLOR_AMARILLO, fill_type="solid")
            
            row += 1
        
        # Ajustar anchos de columna
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Guardar
        wb.save(archivo_salida)
        logger.info(f"✅ Reporte generado: {archivo_salida}")
    
    def cerrar_conexion(self):
        """Cierra conexión a base de datos."""
        if self.connection:
            self.connection.close()
            logger.info("🔌 Conexión a DB2 cerrada")


# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN Y EJEMPLO DE USO
# ═══════════════════════════════════════════════════════════════

def get_db_config() -> Dict[str, str]:
    """
    Obtiene la configuración de la base de datos desde variables de entorno.

    Returns:
        Dict con credenciales de base de datos
    """
    return {
        'host': os.getenv('DB_HOST', 'your_host_here'),
        'port': os.getenv('DB_PORT', '50000'),
        'database': os.getenv('DB_DATABASE', 'your_database_here'),
        'user': os.getenv('DB_USER', 'your_user_here'),
        'password': os.getenv('DB_PASSWORD', 'your_password_here')
    }


def ejemplo_uso():
    """
    Demuestra el uso del sistema de validación automatizada.
    """
    print("=" * 70)
    print("SISTEMA DE VALIDACIÓN AUTOMATIZADA - DEMO")
    print("Desarrollado por: Julián Alexander Juárez Alvarado")
    print("CEDIS Cancún 427 - Tiendas Chedraui")
    print("=" * 70)
    print()
    
    # Configuración de base de datos (cargada de variables de entorno)
    db_config = get_db_config()
    
    # Crear validador
    validador = ValidadorPlanning(db_config)
    
    # Conectar a DB2
    if not validador.conectar_db2():
        print("❌ No se pudo conectar a la base de datos")
        return
    
    # Lista de OCs a validar (ejemplo)
    lista_ocs = [
        '4500123456',
        '4500123457',
        '4500123458',
        '4500123459',
        '4500123460'
    ]
    
    # Validar OCs
    print(f"\n🔍 Validando {len(lista_ocs)} Órdenes de Compra...")
    resultados = validador.validar_multiples_ocs(lista_ocs)
    
    # Mostrar resumen
    print("\n" + "=" * 70)
    print("RESUMEN DE VALIDACIÓN")
    print("=" * 70)
    
    total_alertas_criticas = sum(
        sum(1 for a in r.get('alertas', []) if a['nivel'] == 'CRÍTICA')
        for r in resultados
    )
    
    total_alertas_medias = sum(
        sum(1 for a in r.get('alertas', []) if a['nivel'] == 'MEDIA')
        for r in resultados
    )
    
    print(f"\n📊 OCs Procesadas: {len(resultados)}")
    print(f"🔴 Alertas Críticas: {total_alertas_criticas}")
    print(f"🟡 Alertas Medias: {total_alertas_medias}")
    print(f"\n⏱️  Tiempo de proceso: ~{len(resultados) * 2} segundos")
    print(f"💡 Ahorro vs manual: {150 - len(resultados) * 2} minutos")
    
    # Generar reporte Excel
    archivo_reporte = f"reporte_validacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    validador.generar_reporte_excel(resultados, archivo_reporte)
    
    # Cerrar conexión
    validador.cerrar_conexion()
    
    print("\n" + "=" * 70)
    print("✅ PROCESO COMPLETADO")
    print(f"📄 Reporte disponible en: {archivo_reporte}")
    print("=" * 70)


if __name__ == "__main__":
    ejemplo_uso()
